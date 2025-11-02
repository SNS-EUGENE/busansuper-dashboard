#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook

# File paths
google_sheet_file = '251031 구글시트 기준 재고.xlsx'
system_file = '251101 영업정보시스템 기준 상품목록.xlsx'
output_file = '251101 통합 상품목록(업로드용).xlsx'

print("===== Excel Merge Start =====")

# 1. Read Google Sheet (stock data)
print("\n[1] Reading Google Sheet...")
wb_google = openpyxl.load_workbook(google_sheet_file, data_only=True)
sheet_google = wb_google.active

# Barcode -> Stock mapping
barcode_to_stock = {}
for row_idx, row in enumerate(sheet_google.iter_rows(min_row=3, values_only=True), start=3):
    if not row or len(row) < 11:
        continue

    barcode = str(row[7]).strip() if row[7] else None  # H column: barcode
    initial_stock = row[10]  # K column: initial stock (최초 판매재고수량)

    if barcode and barcode != 'None' and initial_stock is not None:
        try:
            stock_value = int(initial_stock) if isinstance(initial_stock, (int, float)) else 0
            barcode_to_stock[barcode] = stock_value
        except:
            continue

print(f"\n[OK] Extracted {len(barcode_to_stock)} barcodes from Google Sheet")

# 2. Read system file
print("\n[2] Reading system product list...")
wb_system = openpyxl.load_workbook(system_file)
sheet_system = wb_system.active

# 3. Create merged Excel
print("\n[3] Creating merged file...")
wb_output = Workbook()
sheet_output = wb_output.active
sheet_output.title = "Merged"

# Headers
headers = [
    'No.', '상품코드', '상품명', '분류명', '거래처', '주문상품여부',
    '과면세여부', '공급단가', '판매상품여부', '과세구분여부',
    '판매단가', '판매과세여부', '재고여부', '주문관리여부',
    '할인율', '최소주문수량', '외부상품코드', '바코드',
    '사용여부', '진열여부', '자녀', '출력', '매칭키워드사용',
    '주문관리여부2', '비고', '초기재고'
]
sheet_output.append(headers)

# Merge data
matched_count = 0
unmatched_count = 0
unmatched_list = []

for row_idx, row in enumerate(sheet_system.iter_rows(min_row=2, values_only=True), start=2):
    if not row:
        continue

    barcode = str(row[17]).strip() if row[17] else None  # R column: barcode
    product_code = row[1]
    product_name = row[2]

    # Find stock
    initial_stock = 0
    if barcode and barcode in barcode_to_stock:
        initial_stock = barcode_to_stock[barcode]
        matched_count += 1
    else:
        unmatched_count += 1
        unmatched_list.append({
            'code': product_code,
            'name': product_name,
            'barcode': barcode or 'N/A'
        })

    # Create new row
    new_row = list(row) + [initial_stock]
    sheet_output.append(new_row)

# 4. Save file
wb_output.save(output_file)

print(f"\n" + "="*50)
print(f"[SUCCESS] Merge Complete!")
print(f"Output file: {output_file}")
print(f"Statistics:")
print(f"   - Total products: {matched_count + unmatched_count}")
print(f"   - Matched: {matched_count}")
print(f"   - Unmatched: {unmatched_count}")
print(f"="*50)

# Print unmatched products
if unmatched_list:
    print(f"\n[UNMATCHED PRODUCTS - {len(unmatched_list)} items]")
    print("="*80)
    for idx, item in enumerate(unmatched_list, start=1):
        print(f"{idx:3d}. [{item['code']}] {item['name'][:40]:40s} | Barcode: {item['barcode']}")
    print("="*80)
